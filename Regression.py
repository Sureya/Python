import ast
from statsmodels.formula.api import ols
import numpy as np
import scipy as sp
import pandas as pd
import sys,os
import openpyxl as xl
import json
import re 
import logging
import timeit
import time
import MySQLdb

import warnings
warnings.simplefilter(action = "ignore", category = FutureWarning)
warnings.simplefilter(action = "ignore", category = UserWarning)


def comp(list1, list2):
    for val in list1:
        if val in list2:
            return True
    return False

def maintain_benchmark(common,coefs):
    con = MySQLdb.connect (host = "127.0.0.1", port = 3306, user = "Sureya", passwd = "sureya", db = "ecube")
    cursor = con.cursor()

    common = list(set(common))

    print "Connected to Maintainance database  "
    name = '_'.join(coefs.keys())
    flag = checkTableExists(con,name)

    if flag==True:
        print "%s : Already exists"%name
        data = pd.DataFrame()

        for x in common:
            checkdata = pd.read_sql("SELECT * FROM %s"%(x),con=con)
            length  = len(checkdata)-1
            data = data.append(checkdata.ix[length],ignore_index=True)
            index = len(data)-1
            data.loc[index,"Code"] = x.split('_')[0]
        
        varia = list(data.columns.values)
        for x in varia:
            if x == 'index':
                data.drop(x,inplace=True,axis=1)

                

        data.to_sql(name=name,con=con,if_exists='append',flavor='mysql')

    elif flag==False:
        print "%s : New Benchmark created"%name
        data = pd.DataFrame()
        for x in common:
            checkdata = pd.read_sql("SELECT * FROM %s"%(x),con=con)
            length  = len(checkdata)-1
            data =data.append(checkdata.ix[length],ignore_index=True)
            index = len(data)-1
            data.loc[index,"Code"] = x.split('_')[0]

        data.to_sql(name=name,con=con,if_exists='replace',flavor='mysql')


def benchmark(name,coefs):
    con = MySQLdb.connect (host = "127.0.0.1", port = 3306, user = "Sureya", passwd = "sureya", db = "ecube")
    cursor = con.cursor()

    print "Connected to Benchmarking database  "

    flag = checkTableExists(con,name)

    if flag ==True:
        checkdata = pd.read_sql("SELECT * FROM %s"%(name),con=con)
        columns = list(coefs.keys())
        #print columns 
        finals = {}

        for x in columns:
            x= "'"+x+"'"
            cursor.execute(" SELECT DISTINCT TABLE_NAME FROM INFORMATION_SCHEMA.COLUMNS  WHERE COLUMN_NAME =%s  AND TABLE_SCHEMA='ecube';"%(x))
            rows = cursor.fetchall()
            if len(rows)>0:
                finals[x] = rows
            elif len(rows)==0:
                print "No matching data available."

        
        for x in finals:
            finals[x] = list(finals[x])

        new = {}

        for x in finals:
            temp =[]
            for y in finals[x]:
                temp.append(''.join(y))
            new[x[1:-1]] = temp

        
        
        
        vals = new.values()
        uniq = set(vals[0])

        for lst in vals[1:]:
            uniq.intersection_update(lst)

        uniq = list(uniq)
        
        return uniq


    elif flag == False:
        error("Sql Error","Comparision table doesn't exist")
        empty = []
        return empty

def cal_avgcount(input_xl,dep,coefs):
    data  =  pd.read_csv(input_xl)
    quit_flag = False   
    global variables
    variables = list(data.columns.values)
    
    # Check for Date column
    if 'Date' in variables:
        data.drop('Date',inplace=True,axis=1)
        variables.remove("Date")
    
    #check for Dependent variable.
    if dep in variables:
        variables.remove(dep)
        columns_no = len(variables)
        row_no = len(data)

    else:
        error("Error","Dependent variable  Not found in data set.")
        print json.dumps(errors)
        sys.exit()
    
    le = len(variables)
    le = le/2
    lei = le
    
    if len(variables) % 2 != 0:
        error("Error","Odd number of Independent variables found for Spinning Regression.")
        quit_flag = True
    
    if quit_flag==True:
        print json.dumps(errors)
        sys.exit()
    
    for i in range(0,le):
        prod  =  variables[i].split('_')[0] 
        string = str(prod)
        data[string] =  data[variables[i]] * data[variables[lei]]
        lei = lei+1
    
    new = data.copy(deep=True)

    data.drop(variables,inplace=True,axis=1)
    data.drop(dep,inplace=True,axis=1)

    var = list(data.columns.values)
    new.drop(var,inplace=True,axis=1)
    new.drop(dep,inplace=True,axis=1)

    avgs = {}

    for x in var:
        avgs[x+" Avg_count"] = data[x].sum() / new[x+'_prod'].sum()

    return avgs



def checkTableExists(dbcon, tablename):
    dbcur = dbcon.cursor()
    dbcur.execute("""
        SELECT COUNT(*)
        FROM information_schema.tables
        WHERE table_name = '{0}'
        """.format(tablename.replace('\'', '\'\'')))
    if dbcur.fetchone()[0] == 1:
        dbcur.close()
        return True

    dbcur.close()
    return False

def putdata(input_xl,dep,coefs,avgs):

    con = MySQLdb.connect (host = "127.0.0.1", port = 3306, user = "Sureya", passwd = "sureya", db = "ecube")
    cursor = con.cursor()

    print "Connected to localhost database  "

    name = ""

    if '/' in input_xl:
        name = input_xl.split('/')[-1].split('.csv')[0]
        

    if '//' in input_xl:
        name = input_xl.split('//')[-1].split('.csv')[0]
        


    if '\\' in input_xl:
        name = input_xl.split('\\')[-1].split('.csv')[0]

    name =  name.lower()
    flag = checkTableExists(con,name)   



    if flag == True:
        checkdata = pd.read_sql("SELECT * FROM %s"%(name),con=con)
        
        check_columns = list(checkdata.columns.values)
        
        if 'Date' in check_columns:
            check_columns.remove("Date")
        
        if comp(list(coefs.keys()),check_columns) == True:
            check_columns = list(coefs.keys())
        else:
            print "Not present"
        

        data = pd.read_csv(input_xl)
        columns = list(coefs.keys())

        compare = cmp(check_columns,columns)

        if compare == 0:

            df = pd.DataFrame()
            loctime = time.asctime( time.localtime(time.time()) )
            for x in columns:
                df.loc[loctime,x] = coefs[x]
                
            for x in columns:
                df.loc[loctime,x+' Avg_count'] = avgs[x+' Avg_count']

            for x in columns:
                df.loc[loctime,x+" Ukg index"] = df.loc[loctime,x] / df.loc[loctime,x+' Avg_count']


           

            df.to_sql(name=name,con=con,if_exists='append',index_label='Date',flavor='mysql')
            print "Existing table updated with regression details"

        if compare !=0:
            print (check_columns,columns)
            error("Fatal Error","Mismatch in SQL table variables and Data variables")
            print errors
            sys.exit()

    if flag == False:
        data = pd.read_csv(input_xl)
        columns = list(coefs.keys())
        df = pd.DataFrame()
        loctime = time.asctime( time.localtime(time.time()) )
        
        for x in columns:
            df.loc[loctime,x] = coefs[x]
                
            
        for x in columns:
            df.loc[loctime,x+' Avg_count'] = avgs[x+' Avg_count']

        for x in columns:
            df.loc[loctime,x+" Ukg index"] = df.loc[loctime,x] / df.loc[loctime,x+' Avg_count']


        df.to_sql(name=name,con=con,if_exists='replace',index_label='Date',flavor='mysql')
        print "New table updated with regression details"


def error(type,value):

    errors[type] = value


def yes_reg(input_xl,dep):
    
    global iteration
    global row_no
    global columns_no
    global deviated_numbers

    deviated_numbers = 0
    iteration = 1
    data  =  pd.read_csv(input_xl)
    writer = pd.ExcelWriter('generic_output.xlsx')

    fil_name = "generic_equation.txt"
    fil= open(fil_name,'w') 
    
    fi_name = "generic_summary.txt"
    fi= open(fi_name,'w') 
   
    
    quit_flag = False   
    global variables
    variables = list(data.columns.values)
    
    # Check for Date column
    if 'Date' in variables:
        data.drop('Date',inplace=True,axis=1)
        variables.remove("Date")
    
    #check for Dependent variable.
    if dep in variables:
        variables.remove(dep)
        columns_no = len(variables)
        row_no = len(data)

    else:
        error("Error","Dependent variable  Not found in data set.")
        print json.dumps(errors)
        sys.exit()
    
    le = len(variables)
    le = le/2
    lei = le
    
    if len(variables) % 2 != 0:
        error("Error","Odd number of Independent variables found for Spinning Regression.")
        quit_flag = True
    
    if quit_flag==True:
        print json.dumps(errors)
        sys.exit()
    
    for i in range(0,le):
        prod  =  variables[i].split('_')[0] 
        string = str(prod)
        data[string] =  data[variables[i]] * data[variables[lei]]
        lei = lei+1
    
    data.drop(variables,inplace=True,axis=1)
    var = list(data.columns.values)
    var.remove(dep)
    leng = len(var)
    variables = var
    

   
    if leng >1:
        data['TP'] = 0 
      
        
        for x in var:
            
            data['TP'] += data[x]
        
        
        variables = list(data.columns.values)    
        variables.remove(dep)
        
        credentials = {}
        
        for values in variables:
            credentials[values] = {}
        
        for value in variables:
            credentials[value]['high'] = data[value].mean() + data[value].std()
            credentials[value]['low'] = data[value].mean() - data[value].std()
            credentials[value]['mean'] = data[value].mean()
            if len(data[value].mode())<1 :
                credentials[value]['mode'] = 0
            else:
                credentials[value]['mode'] = max(data[value].mode())
        
        for value in variables :
            credentials[value]['high'] = credentials[value]['high'] + credentials[value]['high'] *0.05 
            credentials[value]['low'] = credentials[value]['low'] - credentials[value]['low'] *0.05
            
        for value in variables :
            credentials[value]['high'] = int (credentials[value]['high'])
            credentials[value]['low'] = int(credentials[value]['low'])
            credentials[value]['mean'] = int(credentials[value]['mean'])
            credentials[value]['mode'] = int (credentials[value]['mode'])
        
        tp = {}
        tp['high'] = credentials['TP']['high']
        tp['low'] = credentials['TP']['low']

        ssum = data.describe()
        lis = ["min","25%","50%","75%","max","count"]
        ssum = ssum.drop(lis)
        
        
        ssum.loc["high"] = ssum.loc["mean"] + ssum.loc["std"]   
        ssum.loc["low"] = ssum.loc["mean"] - ssum.loc["std"]
        lis = ["mean","std"]
        ssum = ssum.drop(lis)
        
        ssum.loc['high'] = (ssum.loc['high'])*1.05 
        ssum.loc['low'] = (ssum.loc['low'])*0.95
        
        ssum.to_excel(writer,"Data_Statistics")
        
        core = pd.DataFrame()
        core = data.corr()
        core.to_excel(writer,"Co-relation")
        
        variables.remove("TP")
        data = data[(data['TP']>tp['low'])& (data['TP']<tp['high'])]
        work = []
        num =[]
        
        for value in variables:
         work = ((np.where((data[value] > credentials[value]['high']) | (data[value] < credentials[value]['low']))))
         leng = len(work)
         for x in range(0,leng):
                num.append(work[x])
                
        
        num_lens = []
        for x in num :
            num_lens.append(len(x))
        
        indd =0
        
        for x in range(0,len(num_lens)-1):
            if num_lens[x] > num_lens[x+1]:
                indd = x
            else:
                indd = x+1
                
        test = num[indd]
        
        result = num[0]
        for y in range(0,len(num)-1):
            for x in range(1,len(num)):
                result = np.intersect1d(result,num[x])
                
        res = result.tolist()
        
        data = data.drop(data.index[[res]])
            
        string = dep+"~"

        for values in variables:
            string = string+values+'+'
        
        string = string[:-1]
        string = string+"-1"    
        m1 = ols(string, data).fit() 
        f_value = m1.fvalue
        p_values = m1.pvalues
        r_sqr = m1.rsquared
        coef = m1.params
        Result_str = ""
       
        pflag = 0
        mflag = 0
        
        fi.write(str(m1.summary()))
        fi.close()

        dfr = pd.DataFrame()
        
        for x in variables:
            dfr.loc[x,"p-values"] = p_values[x]
            if p_values[x]>0.05:
                dfr.loc[x,"Comments"] = "Greater than 0.05 "
            else:
                dfr.loc[x,"Comments"] = "Less than 0.05"
        
        dfr.to_excel(writer,"Regression P-values")
        
        dfp = pd.DataFrame()
        for x in variables:
            dfp.loc[x,"Co-efficient"] = coef[x]
        
        dfp.to_excel(writer,"Regression Coefficients")
        
        drop_list = []
        
        for value in variables:
            if p_values[value] >0.05 :
                drop_list.append(value)
                
                pflag = 1
                mflag =1
        

        drop_list.append('TP')
        
        if pflag == 1:

            writer.save()
            data.drop(drop_list,inplace=True,axis=1)
            ret = new_yesreg(data,dep)
        
            
        if pflag!=1:

            if r_sqr <0.98:
                error("Warning","R square value less than 0.98")

            for value in variables:
                Result_str += str(float(coef[value])) +value+'+'
            
            Result_str = Result_str[:-1]
            #print "Final Equation of Regression Analysis is :%s",Result_str
            fil.write(Result_str+'\n')
            #fil.write(coef + '\n'+ '\n'+ '\n')
            fil.close()

            a = Result_str
            a = a.split('+')
            rem= {}
            for x in a :
                r = re.compile("([0-9]+)([.]*)([0-9]*)([a-zA-Z_]+)")
                m = r.match(x)
                if m :
                   value = (float(str(str (m.group(1)) + '.' +str(m.group(3))) ))
                   
                   key = str(m.group(4))
                   rem[key] = value
            ret = rem
            coefs = rem.values()
            coefs = [round(float(i),2) for i in coefs] 

            i=0
            data['Expected '+dep] = 0.0
            for x in variables:
                data['Expected '+dep] += data[x] * coef[x] 
                i= i+1
            data['Deviation%'] = ((data[dep] - data['Expected '+dep])/ (data['Expected '+dep]))*100
            deviated  = data[(data['Deviation%']>5.0)| (data['Deviation%']<-5.0) ]
            deviated.to_excel(writer,"Post-Regression Deviations")
            deviated_numbers = len(deviated)
            writer.save() 

        return ret

    elif leng ==1:

        #print "Enetring 1 column condition"
        column = str (var[0])
        #print column
        
        high = max(data[column])
        low = min(data[column])
        high = high *1.05
        low = low *0.95

        high= int(high)
        low = int(low)
        #print high, low
        ssum = data.describe()
        lis = ["min","25%","50%","75%","max","count"]
        ssum = ssum.drop(lis)
        
        
        ssum.loc["high"] = ssum.loc["mean"] + ssum.loc["std"]   
        ssum.loc["low"] = ssum.loc["mean"] - ssum.loc["std"]
        lis = ["mean","std"]
        ssum = ssum.drop(lis)
        
        ssum.loc['high'] = (ssum.loc['high'])*1.05 
        ssum.loc['low'] = (ssum.loc['low'])*0.95
        
        ssum.to_excel(writer,"Data_Statistics")
        
        core = pd.DataFrame()
        core = data.corr()
        core.to_excel(writer,"Co-relation")
        
        ind = data[(data[column]<low)|(data[column]>high)].index
        ind = ind.tolist()
        data = data.drop(data.index[[ind]])
        str_pass = dep+'~'+ column +'-1'
        m1 = ols(str_pass, data).fit() 
        f_value = m1.fvalue
        p_values = m1.pvalues
        
        fi.write(str(m1.summary()))
        fi.close()
        
        r_sqr = m1.rsquared
        coef = m1.params
        Result_str = ""

        for value in variables:
            Result_str += str(coef[value]) +value+'+'
        
        Result_str = Result_str[:-1]
        fil.write(Result_str)
        fil.close()

        dfr = pd.DataFrame()
        for x in variables:
            dfr.loc[x,"p-values"] = p_values[x]
            if p_values[x]>0.05:
                dfr.loc[x,"Comments"] = "Greater than 0.05 "
            else:
                dfr.loc[x,"Comments"] = "Less than 0.05"
        
        dfr.to_excel(writer,"Regression P-values")
        
        dfp = pd.DataFrame()

        for x in variables:
            dfp.loc[x,"Co-efficient"] = coef[x]
        
        dfp.to_excel(writer,"Regression Coefficients")

        rem = {}
        rem[column] = coef[column]
        ret = rem

        coefs = rem.values()
        coefs = [round(float(i),2) for i in coefs]
        i=0
        data['Expected '+dep] = 0.0
        for x in variables:
            data['Expected '+dep] += data[x] * coef[x] 
            i= i+1

        data['Deviation%'] = ((data[dep] - data['Expected '+dep])/ (data['Expected '+dep]))*100

        deviated  = data[(data['Deviation%']>5.0)| (data['Deviation%']<-5.0) ]

        deviated.to_excel(writer,"Post-Regression Deviations")
        deviated_numbers = len(deviated)
        writer.save() 
        return ret
        
    elif leng <= 0:
        error("Warning","Model doesn't fit, No co-relation found between Independent variables and Dependent variables ")
       
        sys.exit()

 
def new_yesreg(df,dep):

    
    file_name = 'generic_output.xlsx'

    fil_name = "generic_equation.txt"
    fil= open(fil_name,'w') 
    
    fi_name = "generic_summary.txt"
    fi= open(fi_name,'w') 
   
    
    wb = xl.load_workbook(file_name)
    x_sheet =  wb.get_sheet_names()
    x_sheet.remove("Regression Coefficients")
    if "Post-Regression Deviations" in x_sheet:
        x_sheet.remove("Post-Regression Deviations")

    df_lis = []
    
    for x in x_sheet:
        xl1  =  pd.ExcelFile(file_name)
        data = xl1.parse(x)
        df_lis.append(data)

    
    writer = pd.ExcelWriter(file_name)
    
    for x in range(0,len(x_sheet)):
        df_lis[x].to_excel(writer,x_sheet[x])



    var1 = list(df.columns.values)
    
    if dep in var1:
        var1.remove(dep)
        columns_no = len(var1)
        row_no = len(df)
    else:
        error("Error","Dependent variable not found")
        sys.exit()
    
    leng = len(var1)

    if leng>1:
        df['TP'] = 0 


        
        for x in var1:
            
            df['TP'] += df[x]
        
        
        var = list(df.columns.values)    
        var.remove(dep)
        
        credentials = {}
        
        for values in var:
            credentials[values] = {}
        
        for value in var:
            credentials[value]['high'] = df[value].mean() + df[value].std()
            credentials[value]['low'] = df[value].mean() - df[value].std()
            credentials[value]['mean'] = df[value].mean()
            if len(df[value].mode())<1 :
                credentials[value]['mode'] = 0
            else:
                credentials[value]['mode'] = max(df[value].mode())
        
        for value in var :
            credentials[value]['high'] = credentials[value]['high'] + credentials[value]['high'] *0.05 
            credentials[value]['low'] = credentials[value]['low'] - credentials[value]['low'] *0.05
            
        for value in var :
            credentials[value]['high'] = int (credentials[value]['high'])
            credentials[value]['low'] = int(credentials[value]['low'])
            credentials[value]['mean'] = int(credentials[value]['mean'])
            credentials[value]['mode'] = int (credentials[value]['mode'])
        
        tp = {}
        tp['high'] = credentials['TP']['high']
        tp['low'] = credentials['TP']['low']
        
        
        var.remove("TP")
        df = df[(df['TP']>tp['low'])& (df['TP']<tp['high'])]
        work = []
        num =[]
        
        for value in var:
         work = ((np.where((df[value] > credentials[value]['high']) | (df[value] < credentials[value]['low']))))
         leng = len(work)
         for x in range(0,leng):
                num.append(work[x])
                
        
        num_lens = []
        for x in num :
            num_lens.append(len(x))
        
        indd =0
        
        for x in range(0,len(num_lens)-1):
            if num_lens[x] > num_lens[x+1]:
                indd = x
            else:
                indd = x+1
                
        test = num[indd]
        
        result = num[0]
        for y in range(0,len(num)-1):
            for x in range(1,len(num)):
                result = np.intersect1d(result,num[x])
                
        res = result.tolist()
        
        df = df.drop(df.index[[res]])
            
        string = dep+"~"

        for values in var:
            string = string+values+'+'
                
        string = string[:-1]
        string = string+"-1"
        m1 = ols(string, df).fit() 
        f_value = m1.fvalue

        fi.write(str(m1.summary()))
        fi.close()

        p_values = m1.pvalues
        r_sqr = m1.rsquared
        coef = m1.params

        

        Result_str = ""
        pflag = 0
        mflag = 0
        drop_list = []

        dfp = pd.DataFrame()
        for x in var:
            dfp.loc[x,"Co-efficient"] = coef[x]
        
        dfp.to_excel(writer,"Regression Coefficients")
        
        drop_list = []
        
        for value in var:
            if p_values[value] >0.05 :
                drop_list.append(value)
                pflag = 1
                mflag =1
        
        drop_list.append('TP')
        
        if pflag == 1:
            df.drop(drop_list,inplace=True,axis=1)
            new_yesreg(df,dep)
        
            
        if pflag!=1:
            string_Add = ""
            for val in variables:
                if val not in var:
                    string_Add += '0'+ str(val) + '+'
            string_Add = string_Add[:-1] 

            for value in var:
                Result_str += str(float(coef[value])) +value+'+'
            
            Result_str = Result_str[:-1]
            Result_str = Result_str +'+'+ string_Add

            fil.write(Result_str+'\n')
            #fil.write(coef + '\n'+ '\n'+ '\n')
            fil.close()

            a = Result_str
            a = a.split('+')
            rem= {}
            for x in a :
                r = re.compile("([0-9]+)([.]*)([0-9]*)([a-zA-Z_]+)")
                m = r.match(x)
                if m :
                   value = (float(str(str (m.group(1)) + '.' +str(m.group(3))) ))
                   
                   key = str(m.group(4))
                   rem[key] = value

            coefs = rem.values()
            coefs = [round(float(i),2) for i in coefs] 

            i=0
            df['Expected '+dep] = 0.0
            
            for x in var:
                df['Expected '+dep] += df[x] * coef[x] 
                i= i+1

            df['Deviation%'] = ((df[dep] - df['Expected '+dep])/ (df['Expected '+dep]))*100

            deviated  = df[(df['Deviation%']>5.0)| (df['Deviation%']<-5.0) ]

            deviated.to_excel(writer,"Post-Regression Deviations")
            deviated_numbers = len(deviated)
            writer.save() 

            return rem

    elif leng ==1:
        column = str(df[var1[0]])
        high = max(df[column])
        low = min(df[column])
        high = high *1.05
        low = low *0.95

        high= int(high)
        low = int(low)

        ind = df[(df[column]<low)|(df[column]>high)].index
        ind = ind.tolist()
        df = df.drop(df.index[[ind]])
        str_pass = dep+'~'+ column +'-1'
        m1 = ols(str_pass, data).fit() 
        f_value = m1.fvalue
        p_values = m1.pvalues
        
        fi.write(str(m1.summary()))
        fi.close()

        r_sqr = m1.rsquared
        coef = m1.params
        rem = {}
        rem[column] = coef[column]

        Result_str = ""
        for x in rem:
            Result_str = rem[x]+x+'+'
        Result_str = Result_str[:-1]
        coefs = rem.values()
        coefs = [round(float(i),2) for i in coefs] 

        i=0
        df['Expected '+dep] = 0.0
        for x in var1:
            df['Expected '+dep] += df[x] * coef[x] 
            i= i+1

        df['Deviation%'] = ((df[dep] - df['Expected '+dep])/ (df['Expected '+dep]))*100

        deviated  = df[(df['Deviation%']>5.0)| (df['Deviation%']<-5.0) ]

        deviated.to_excel(writer,"Post-Regression Deviations")
        deviated_numbers = len(deviated)
        writer.save() 
        return rem

    elif leng<=0:
        error("Warning","Model doesn't fit, No co-relation found between Independent variables and Dependent variables ")
       
        sys.exit()

def no_reg(input_xl,dep):
    
    global iteration
    iteration = 1

    global deviated_numbers

    deviated_numbers = 0

    global row_no
    global columns_no

    
    data  =  pd.read_csv(input_xl)
    global variables
   
    variables = list(data.columns.values)    
    
    writer = pd.ExcelWriter('generic_output.xlsx')
    
    fil_name = "generic_equation.txt"
    fil= open(fil_name,'w') 
    
    fi_name = "generic_summary.txt"
    fi= open(fi_name,'w') 
    


    if 'Date' in variables:
        data.drop('Date',axis=1,inplace=True)
        variables.remove("Date")
        
    if dep in variables:
        variables.remove(dep)
        columns_no = len(variables)
        row_no = len(data)
    else:
        error("Error","Dependent variable not found.")
        sys.exit()

    ssum = data.describe()
    lis = ["min","25%","50%","75%","max","count"]
    ssum = ssum.drop(lis)
    
    
    ssum.loc["high"] = ssum.loc["mean"] + ssum.loc["std"]   
    ssum.loc["low"] = ssum.loc["mean"] - ssum.loc["std"]
    lis = ["mean","std"]
    ssum = ssum.drop(lis)
    
    ssum.loc['high'] = (ssum.loc['high'])*1.05 
    ssum.loc['low'] = (ssum.loc['low'])*0.95
    
    ssum.to_excel(writer,"Data_Statistics")
    
    core = pd.DataFrame()
    core = data.corr()
    core.to_excel(writer,"Co-relation")
    
    string = dep+"~"
    
    for values in variables:
        string = string+values+'+'
        
    string = string[:-1]
    string = string+"-1"
    
    m1 = ols(string, data).fit() 
    
    f_value = m1.fvalue
    p_values = m1.pvalues
    r_sqr = m1.rsquared
    coef = m1.params
    
    Result_str = ""
    
    pflag = 0
    
    fi.write(str(m1.summary()))
    fi.close()

    dfr = pd.DataFrame()
    for x in variables:
        dfr.loc[x,"p-values"] = p_values[x]
        if p_values[x]>0.05:
            dfr.loc[x,"Comments"] = "Greater than 0.05 "
        else:
            dfr.loc[x,"Comments"] = "Less than 0.05"
    
    dfr.to_excel(writer,"Regression P-values")
    
    dfp = pd.DataFrame()
    for x in variables:
        dfp.loc[x,"Co-efficient"] = coef[x]
    
    dfp.to_excel(writer,"Regression Coefficients")
    
    

        
    drop_list = []
    for value in variables:
        if p_values[value] >0.05 :
            drop_list.append(value)
            pflag = 1
    
    if pflag == 1:
        writer.save()
        data.drop(drop_list,axis=1)
        #print "Calling noloop"
        #print variables
        new_noreg(data,dep)
    
    if pflag!=1 :
        for value in variables :
            Result_str += str(coef[value]) +value+'+'
    
        Result_str = Result_str[:-1]
        
        fil.write(Result_str)
        fil.close()

        coefs = []
        for x in variables:
            coefs.append(coef[x])
        coefs = [round(float(i),2) for i in coefs] 

        
        i=0
        data['Expected '+dep] = 0.0
        
        for x in variables:
            data['Expected '+dep] += data[x] * coef[x] 
            i= i+1

        data['Deviation%'] = ((data[dep] - data['Expected '+dep])/ (data['Expected '+dep]))*100

        deviated  = data[(data['Deviation%']>5.0)| (data['Deviation%']<-5.0) ]

        deviated.to_excel(writer,"Post-Regression Deviations")
        deviated_numbers = len(deviated)

        writer.save()    
 
def new_noreg(df,dep):
    
    global iteration
    
   # print "Coming inside noloop"
    iteration = iteration +1
    file_name = 'generic_output.xlsx'
    writer = pd.ExcelWriter(file_name)
    
    
    wb = xl.load_workbook(file_name)
    x_sheet =  wb.get_sheet_names()
    
    x_sheet.remove("Regression Coefficients")

    if "Post-Regression Deviations" in x_sheet:
        x_sheet.remove("Post-Regression Deviations")


    df_lis = []
    
    for x in x_sheet:
        xl1  =  pd.ExcelFile(file_name)
        data = xl1.parse(x)
        df_lis.append(data)

    
    for x in range(0,len(x_sheet)):
        df_lis[x].to_excel(writer,x_sheet[x])

    var = list(df.columns.values)   
    #print var
    fil_name = "generic_equation.txt"
    fil= open(fil_name,'w') 
    
    fi_name = "generic_summary.txt"
    fi= open(fi_name,'w') 

    if dep in var:
        var.remove(dep)
        columns_no = len(var)
        row_no = len(df)

    else:
       print("Error","Dependent variable not found.")

       sys.exit()

    string = dep+"~"
    string_Add = ""
    
    
    for values in var:
        string = string+values+'+'
        
    
    string = string[:-1]
    string = string+"-1"
    
    m1 = ols(string, df).fit() 
    
    f_value = m1.fvalue
    p_values = m1.pvalues
    r_sqr = m1.rsquared
    coef = m1.params
    
    Result_str = ""
    
    pflag = 0
    
    # print "visited Fi"
    fi.write(str(m1.summary()))
    fi.close()

    dfp = pd.DataFrame()
    for x in var:
        dfp.loc[x,"Co-efficient"] = coef[x]
    
    dfp.to_excel(writer,"Regression Coefficients")
    
    drop_list = []
    for value in var:
        if p_values[value] >0.05 :
            drop_list.append(value)
            pflag = 1
    
    if pflag == 1:
        
        writer.save()
        df.drop(drop_list,axis=1,inplace=True)
        new_noreg(df,dep)
        
    if pflag!=1 :
        for value in var :
            Result_str += str(coef[value]) +value+'+'
    
        Result_str = Result_str[:-1]
        
        for val in variables:
            if val not in var:
                string_Add += '0'+ str(val) + '+'
        string_Add = string_Add[:-1] 
    
        Result_str = Result_str +'+'+ string_Add
        #print "visited File"
        fil.write(Result_str)
        fil.close()

        #df = df.drop("TP",axis=1)
        coefs = []
        for x in var:
            coefs.append(coef[x])

        coefs = [round(float(i),2) for i in coefs] 
        
        i=0
        df['Expected '+dep] = 0.0
        
        for x in var:
            df['Expected '+dep] += df[x] * coef[x] 
            i= i+1

        df['Deviation%'] = ((df[dep] - df['Expected '+dep])/ (df['Expected '+dep]))*100

        deviated  = df[(df['Deviation%']>5.0)| (df['Deviation%']<-5.0) ]

        deviated.to_excel(writer,"Post-Regression Deviations")
        deviated_numbers = len(deviated)

        writer.save()



        print "Saved"
    
def analyze(input_xl):
    
    file_name = 'generic_output.xlsx'
    wb = xl.load_workbook(file_name)
    x_sheet =  wb.get_sheet_names()

    df_lis = []
    
    for x in x_sheet:
        xl1  =  pd.ExcelFile(file_name)
        data = xl1.parse(x)
        df_lis.append(data)

    writer = pd.ExcelWriter(file_name)
    
    for x in range(0,len(x_sheet)):
        df_lis[x].to_excel(writer,x_sheet[x])

    
    index = input("Enetr the Index to which similar data to be displayed: ")
    index = int(index)
    data  =  pd.read_csv(input_xl)
    
    #print len(data)
    
    valss = []
    hi_vals = []
    lo_vals = []

    for x in data.ix[index]:
        valss.append(x)

    valss.remove(valss[0])
    valss.remove(valss[-1])
    #print valss

    for x in valss:
        hi_vals.append(x*1.05)
        lo_vals.append(x*0.95)
        

    final_lis = []
    lim = len(data)

    # Master filter.
    for x in range(0,lim):
        vals = []
        for cal in data.ix[x]:
            vals.append(cal)
        
        vals.remove(vals[0])
        vals.remove(vals[-1])

        truth = []
        if len(vals)==len(lo_vals):
            for i in range(len(hi_vals)):
                if ((vals[i]> lo_vals[i]) and vals[i]<hi_vals[i]):
                            truth.append(True)
                            
                else:
                    truth.append(False)
                    
            if (False in truth):
                something=1
            else:
                final_lis.append(x)
        else:
            print "Length mismatch"
            
    result = data.ix[final_lis]
    if len(result)>0:
        result.to_excel(writer,"Similar data to index of  "+str(index))
        writer.save()
    elif len(result) ==0:
        print "No similar data set found within '5%' deviation"

    

if __name__ == "__main__":

    start = timeit.default_timer()
    global errors
    errors = {}

    input_xl = sys.argv[1]
    deviated_numbers =0
    row_no =0
    columns_no = 0

    if '/' in input_xl:
        dep = input_xl.split('/')[-1].split('.csv')[0].split('_')[1]
        

    if '//' in input_xl:
        dep = input_xl.split('//')[-1].split('.csv')[0].split('_')[1]
        


    if '\\' in input_xl:
        dep = input_xl.split('\\')[-1].split('.csv')[0].split('_')[1]

    

    if sys.argv[2] =="yes":
        res = yes_reg(input_xl,dep)
    
    elif sys.argv[2] =="no":
        res = no_reg(input_xl,dep)
    
   

    opt = raw_input("Do you want to analyze the data set? Enter y to proceed and any other button to exit")
    if opt.lower() == "y":
        analyze(input_xl)
    
    avg = cal_avgcount(input_xl,dep,res)

    
    putdata(input_xl,dep,res,avg)

    if '/' in input_xl:
        as_name = input_xl.split('/')[-1].split('.csv')[0]

    if '//' in input_xl:
        as_name = input_xl.split('//')[-1].split('.csv')[0]
        


    if '\\' in input_xl:
        as_name = input_xl.split('\\')[-1].split('.csv')[0]

    common = benchmark(as_name,res)

    
    if len(common)>1:
        maintain_benchmark(common,res)



    
    if len(errors)==0:
        error("Succesful","No errors found")
    
    print errors
    
    #os.startfile("generic_output.xlsx")

